import os
import datetime
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file, Response
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from flask_wtf import FlaskForm
import uuid
import csv
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None

from app import db
from models import User, Attendance, TreeData
from utils import admin_required

admin_bp = Blueprint('admin', __name__)


@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    # Create a form for CSRF protection
    form = FlaskForm()
    
    # Only count employees (exclude admins)
    total_users = User.query.filter_by(role='employee').count()
    
    # Today's statistics (only employees)
    today = datetime.datetime.now().date()
    today_start = datetime.datetime.combine(today, datetime.time.min)
    today_end = datetime.datetime.combine(today, datetime.time.max)
    today_attendance = db.session.query(Attendance).join(User).filter(
        Attendance.check_in_time >= today_start,
        Attendance.check_in_time <= today_end,
        User.role == 'employee'
    ).count()
    
    # Get recent attendance records (only employees)
    recent_attendance = db.session.query(
        Attendance, User
    ).join(
        User, Attendance.user_id == User.id
    ).filter(
        User.role == 'employee'
    ).order_by(
        Attendance.check_in_time.desc()
    ).limit(10).all()
    
    # Monthly statistics (only employees)
    current_month = datetime.datetime.now().month
    current_year = datetime.datetime.now().year
    
    monthly_stats = db.session.query(
        db.func.extract('day', Attendance.check_in_time).label('day'),
        db.func.count(Attendance.id).label('count')
    ).join(User).filter(
        db.extract('month', Attendance.check_in_time) == current_month,
        db.extract('year', Attendance.check_in_time) == current_year,
        User.role == 'employee'
    ).group_by('day').all()
    
    # Format data for chart
    days = [int(d[0]) for d in monthly_stats]
    counts = [d[1] for d in monthly_stats]
    
    # Get status distribution (only employees)
    current_month_start = datetime.datetime(current_year, current_month, 1)
    next_month = current_month + 1 if current_month < 12 else 1
    next_year = current_year if current_month < 12 else current_year + 1
    current_month_end = datetime.datetime(next_year, next_month, 1) - datetime.timedelta(days=1)
    current_month_end = datetime.datetime.combine(current_month_end, datetime.time.max)
    
    present_count = db.session.query(Attendance).join(User).filter(
        Attendance.check_in_time >= current_month_start,
        Attendance.check_in_time <= current_month_end,
        Attendance.status == 'present',
        User.role == 'employee'
    ).count()
    
    late_count = db.session.query(Attendance).join(User).filter(
        Attendance.check_in_time >= current_month_start,
        Attendance.check_in_time <= current_month_end,
        Attendance.status == 'late',
        User.role == 'employee'
    ).count()
    
    absent_count = db.session.query(Attendance).join(User).filter(
        Attendance.check_in_time >= current_month_start,
        Attendance.check_in_time <= current_month_end,
        Attendance.status == 'absent',
        User.role == 'employee'
    ).count()
    
    # Department statistics (only employees with department info)
    department_stats = db.session.query(
        User.department,
        db.func.count(User.id).label('count')
    ).filter(
        User.department != None,
        User.role == 'employee'
    ).group_by(User.department).all()
    
    dept_labels = [dept[0] for dept in department_stats]  # Department names
    dept_counts = [dept[1] for dept in department_stats]  # User counts in each department
    dept_count = len(department_stats)
    
    return render_template(
        'admin/index.html',
        total_users=total_users,
        today_attendance=today_attendance,
        recent_attendance=recent_attendance,
        days=days,
        counts=counts,
        present_count=present_count,
        late_count=late_count,
        absent_count=absent_count,
        dept_count=dept_count,
        dept_labels=dept_labels,
        dept_counts=dept_counts,
        form=form
    )


@admin_bp.route('/users')
@login_required
@admin_required
def users():
    search = request.args.get('search', '')
    role = request.args.get('role', '')
    
    form = FlaskForm()
    
    query = User.query
    
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.full_name.ilike(f'%{search}%'))
        )
    
    if role:
        query = query.filter(User.role == role)
    
    users = query.order_by(User.created_at.desc()).all()
    
    return render_template('admin/users.html', users=users, form=form)


@admin_bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    # Create a form for CSRF protection
    form = FlaskForm()
    
    # Get distinct departments and positions for dropdowns
    departments = db.session.query(User.department).filter(User.department != None).distinct().all()
    departments = [d[0] for d in departments]
    positions = db.session.query(User.position).filter(User.position != None).distinct().all()
    positions = [p[0] for p in positions]
    
    if request.method == 'POST':
        if form.validate_on_submit():
            username = request.form.get('username')
            email = request.form.get('email')
            password = request.form.get('password')
            full_name = request.form.get('full_name')
            role = request.form.get('role')
            department = request.form.get('department')
            position = request.form.get('position')
            
            # Validate username and email
            if User.query.filter_by(username=username).first():
                flash('Username already exists.', 'danger')
                return redirect(url_for('admin.add_user'))
            
            if User.query.filter_by(email=email).first():
                flash('Email already exists.', 'danger')
                return redirect(url_for('admin.add_user'))
            
            # Create new user
            new_user = User(
                username=username,
                email=email,
                full_name=full_name,
                role=role,
                department=department,
                position=position
            )
            new_user.set_password(password)
            
            db.session.add(new_user)
            db.session.commit()
            
            flash('User has been created successfully.', 'success')
            return redirect(url_for('admin.users'))
        else:
            flash('Please fill all required fields correctly.', 'danger')
            return redirect(url_for('admin.add_user'))
    
    return render_template('admin/edit_user.html', user=None, form=form, departments=departments, positions=positions)


@admin_bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Create a form for CSRF protection
    form = FlaskForm()
    
    # Get distinct departments and positions for dropdowns
    departments = db.session.query(User.department).filter(User.department != None).distinct().all()
    departments = [d[0] for d in departments]
    positions = db.session.query(User.position).filter(User.position != None).distinct().all()
    positions = [p[0] for p in positions]
    
    if request.method == 'POST':
        if form.validate_on_submit():
            email = request.form.get('email')
            full_name = request.form.get('full_name')
            role = request.form.get('role')
            department = request.form.get('department')
            position = request.form.get('position')
            password = request.form.get('password')
            
            # Check if email already exists
            if email != user.email and User.query.filter_by(email=email).first():
                flash('Email already exists.', 'danger')
                return redirect(url_for('admin.edit_user', user_id=user_id))
            
            # Update user
            user.email = email
            user.full_name = full_name
            user.role = role
            user.department = department
            user.position = position
            
            if password:
                user.set_password(password)
            
            db.session.commit()
            
            flash('User has been updated successfully.', 'success')
            return redirect(url_for('admin.users'))
        else:
            flash('Please fill all required fields correctly.', 'danger')
            return redirect(url_for('admin.edit_user', user_id=user_id))
    
    return render_template('admin/edit_user.html', user=user, form=form, departments=departments, positions=positions)


@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    form = FlaskForm()
    if not form.validate_on_submit():
        flash('CSRF token validation failed.', 'danger')
        return redirect(url_for('admin.users'))
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.users'))
    
    db.session.delete(user)
    db.session.commit()
    
    flash('User has been deleted successfully.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/attendance')
@login_required
@admin_required
def attendance():
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status = request.args.get('status', '')
    
    # Create a form for CSRF protection
    form = FlaskForm()
    
    # Base query with join
    query = db.session.query(Attendance, User).join(User, Attendance.user_id == User.id)
    
    # Apply filters
    if search:
        query = query.filter(
            (User.username.ilike(f'%{search}%')) |
            (User.email.ilike(f'%{search}%')) |
            (User.full_name.ilike(f'%{search}%'))
        )
    
    if date_from:
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d')
        query = query.filter(Attendance.check_in_time >= date_from_obj)
    
    if date_to:
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d')
        date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
        query = query.filter(Attendance.check_in_time <= date_to_obj)
    
    if status:
        query = query.filter(Attendance.status == status)
    
    # Order by check-in time (newest first)
    attendances = query.order_by(Attendance.check_in_time.desc()).all()
    
    return render_template(
        'admin/attendance.html',
        attendances=attendances,
        search=search,
        date_from=date_from,
        date_to=date_to,
        status=status,
        form=form
    )


@admin_bp.route('/attendance/delete/<int:attendance_id>', methods=['POST'])
@login_required
@admin_required
def delete_attendance(attendance_id):
    form = FlaskForm()
    if not form.validate_on_submit():
        flash('CSRF token validation failed.', 'danger')
        return redirect(url_for('admin.attendance'))
        
    attendance = Attendance.query.get_or_404(attendance_id)
    
    db.session.delete(attendance)
    db.session.commit()
    
    flash('Attendance record has been deleted successfully.', 'success')
    return redirect(url_for('admin.attendance'))


@admin_bp.route('/hierarchical')
@login_required
@admin_required
def hierarchical():
    # The page fetches the tree via AJAX from /admin/hierarchy-data
    users = User.query.order_by(User.full_name).all()
    return render_template('admin/Hierarchical.html', users=users)


def _build_tree():
    roots = TreeData.query.filter_by(parent_id=None).order_by(TreeData.id).all()
    return [root.to_dict() for root in roots]


@admin_bp.route('/hierarchy-data')
@login_required
@admin_required
def hierarchy_data():
    tree = _build_tree()
    return jsonify(tree)


@admin_bp.route('/hierarchy-node', methods=['POST'])
@login_required
@admin_required
def add_hierarchy_node():
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    parent_id = data.get('parent_id')

    if not name:
        return jsonify({'success': False, 'message': 'Name is required.'}), 400

    if parent_id:
        parent = TreeData.query.get(parent_id)
        if not parent:
            return jsonify({'success': False, 'message': 'Parent node not found.'}), 404

    node = TreeData(name=name, parent_id=parent_id)
    db.session.add(node)
    db.session.commit()
    return jsonify({'success': True, 'node': node.to_dict()}), 201


@admin_bp.route('/hierarchy-node/<int:node_id>', methods=['PUT'])
@login_required
@admin_required
def update_hierarchy_node(node_id):
    node = TreeData.query.get_or_404(node_id)
    data = request.get_json() or {}
    name = data.get('name', '').strip()
    parent_id = data.get('parent_id')

    if not name:
        return jsonify({'success': False, 'message': 'Name is required.'}), 400

    if parent_id == node.id:
        return jsonify({'success': False, 'message': 'A node cannot be its own parent.'}), 400

    if parent_id:
        parent = TreeData.query.get(parent_id)
        if not parent:
            return jsonify({'success': False, 'message': 'Parent node not found.'}), 404

    node.name = name
    node.parent_id = parent_id
    db.session.commit()
    return jsonify({'success': True, 'node': node.to_dict()}), 200


@admin_bp.route('/hierarchy-node/<int:node_id>', methods=['DELETE'])
@login_required
@admin_required
def delete_hierarchy_node(node_id):
    node = TreeData.query.get_or_404(node_id)
    db.session.delete(node)
    db.session.commit()
    return jsonify({'success': True}), 200


@admin_bp.route('/hierarchy-template')
@login_required
@admin_required
def hierarchy_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Name', 'Parent_ID'])
    writer.writerow([1, 'A', ''])
    writer.writerow([2, 'B', 1])
    writer.writerow([3, 'C', 1])
    writer.writerow([4, 'D', 2])
    writer.writerow([5, 'E', 2])
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=hierarchy_template.csv'}
    )


@admin_bp.route('/hierarchy-bulk-upload', methods=['POST'])
@login_required
@admin_required
def hierarchy_bulk_upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': 'No file uploaded.'}), 400

    filename = file.filename.lower()
    rows = []

    if filename.endswith('.csv'):
        stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
        reader = csv.DictReader(stream)
        for row in reader:
            rows.append({
                'id': row.get('ID') and row.get('ID').strip(),
                'name': row.get('Name') and row.get('Name').strip(),
                'parent_id': row.get('Parent_ID') and row.get('Parent_ID').strip()
            })
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        if not openpyxl:
            return jsonify({'success': False, 'message': 'openpyxl is required for Excel upload.'}), 500
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {h: i for i, h in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({
                'id': str(row[header_map.get('ID')]).strip() if row[header_map.get('ID')] is not None else '',
                'name': str(row[header_map.get('Name')]).strip() if row[header_map.get('Name')] is not None else '',
                'parent_id': str(row[header_map.get('Parent_ID')]).strip() if row[header_map.get('Parent_ID')] is not None else ''
            })
    else:
        return jsonify({'success': False, 'message': 'Unsupported file type.'}), 400

    import_ids = set()
    for r in rows:
        if not r['id'] or not r['name']:
            return jsonify({'success': False, 'message': 'Each row must have ID and Name.'}), 400
        if r['id'] in import_ids:
            return jsonify({'success': False, 'message': f'Duplicate ID {r["id"]} in file.'}), 400
        import_ids.add(r['id'])

    pending = rows[:]
    new_map = {}
    created = 0

    while pending:
        before = len(pending)
        for r in pending[:]:
            parent_val = r['parent_id']
            parent_db_id = None

            if parent_val not in (None, '', 'None'):
                try:
                    parent_val_int = int(parent_val)
                except ValueError:
                    return jsonify({'success': False, 'message': f'Invalid Parent_ID {parent_val}.'}), 400

                parent_db_id = new_map.get(str(parent_val_int))
                if parent_db_id is None:
                    existing_parent = TreeData.query.get(parent_val_int)
                    if existing_parent:
                        parent_db_id = existing_parent.id

                if parent_db_id is None:
                    continue

            node = TreeData(name=r['name'], parent_id=parent_db_id)
            db.session.add(node)
            db.session.flush()
            new_map[str(r['id'])] = node.id
            pending.remove(r)
            created += 1

        if len(pending) == before:
            return jsonify({'success': False, 'message': 'Could not resolve some Parent_ID values; check ID relationships.'}), 400

    db.session.commit()
    return jsonify({'success': True, 'created': created}), 201


@admin_bp.route('/reports')
@login_required
@admin_required
def reports():
    # Create a form for CSRF protection
    form = FlaskForm()
    
    # Department statistics (only employees)
    dept_stats = db.session.query(
        User.department,
        db.func.count(User.id).label('user_count')
    ).filter(
        User.department != None,
        User.role == 'employee'
    ).group_by(User.department).all()
    
    # Attendance statistics by status (only employees)
    status_stats = db.session.query(
        Attendance.status,
        db.func.count(Attendance.id).label('count')
    ).join(User).filter(
        User.role == 'employee'
    ).group_by(Attendance.status).all()
    
    # Weekly attendance (only employees)
    today = datetime.datetime.now().date()
    start_of_week = today - datetime.timedelta(days=today.weekday())
    
    daily_stats = []
    for i in range(7):
        day = start_of_week + datetime.timedelta(days=i)
        day_start = datetime.datetime.combine(day, datetime.time.min)
        day_end = datetime.datetime.combine(day, datetime.time.max)
        
        count = db.session.query(Attendance).join(User).filter(
            Attendance.check_in_time >= day_start,
            Attendance.check_in_time <= day_end,
            User.role == 'employee'
        ).count()
        
        daily_stats.append({
            'day': day.strftime('%A'),
            'date': day.strftime('%Y-%m-%d'),
            'count': count
        })
    
    return render_template(
        'admin/reports.html',
        dept_stats=dept_stats,
        status_stats=status_stats,
        daily_stats=daily_stats,
        form=form
    )

@admin_bp.route('/export/attendance', methods=['GET'])
@login_required
@admin_required
def export_attendance():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = db.session.query(
        User.id,
        User.username,
        User.full_name,
        User.department,
        Attendance.check_in_time,
        Attendance.check_out_time,
        Attendance.status,
        Attendance.location,
        Attendance.notes
    ).join(User, Attendance.user_id == User.id)
    
    # Apply date filters
    if date_from:
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d')
        query = query.filter(Attendance.check_in_time >= date_from_obj)
    
    if date_to:
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d')
        date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
        query = query.filter(Attendance.check_in_time <= date_to_obj)
    
    # Order by check-in time
    results = query.order_by(Attendance.check_in_time.desc()).all()
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'User ID', 'Username', 'Full Name', 'Department',
        'Check-in Time', 'Check-out Time', 'Status', 'Location', 'Notes'
    ])
    
    # Write data
    for row in results:
        writer.writerow([
            row.id,
            row.username,
            row.full_name,
            row.department,
            row.check_in_time.strftime('%Y-%m-%d %H:%M:%S') if row.check_in_time else '',
            row.check_out_time.strftime('%Y-%m-%d %H:%M:%S') if row.check_out_time else '',
            row.status,
            row.location,
            row.notes
        ])
    
    # Prepare response
    from flask import Response
    output.seek(0)
    
    filename = f"attendance_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
